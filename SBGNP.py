"""
SBGNP - Situation-Based Genetic Network Programming
"""

# ============================================================================
# IMPORTS
# ============================================================================

import GA_tile_world_instructions as tileins
import GA_common_instructions as comins
import variables
import random
import copy
import threading
import time
import openpyxl as xl


# ============================================================================
# GNP NETWORK STRUCTURE CONFIGURATION
# ============================================================================

# Base individual structure for GNP network (PG=5)
# individual nodes start as 1
# Format: [node_type, function, step_cost, connection_genes...]
# node_type: 0=start, 1=judgement, 2=processing
base_individual = [[], [0, 0, 0], [1, tileins.what_exists_front, 1], [1, tileins.what_exists_right, 1], 
[1, tileins.what_exists_left, 1], [1, tileins.what_exists_back, 1], [1, tileins.nearest_tile_direction, 1], 
[1, tileins.second_nearest_tile_direction, 1], [1, tileins.nearest_hole_direction, 1], [1, tileins.what_exists_front, 1], [1, tileins.what_exists_right, 1], 
[1, tileins.what_exists_left, 1], [1, tileins.what_exists_back, 1], [1, tileins.nearest_tile_direction, 1], 
[1, tileins.second_nearest_tile_direction, 1], [1, tileins.nearest_hole_direction, 1], [1, tileins.what_exists_front, 1], [1, tileins.what_exists_right, 1], 
[1, tileins.what_exists_left, 1], [1, tileins.what_exists_back, 1], [1, tileins.nearest_tile_direction, 1], 
[1, tileins.second_nearest_tile_direction, 1], [1, tileins.nearest_hole_direction, 1], [1, tileins.what_exists_front, 1], [1, tileins.what_exists_right, 1], 
[1, tileins.what_exists_left, 1], [1, tileins.what_exists_back, 1], [1, tileins.nearest_tile_direction, 1], 
[1, tileins.second_nearest_tile_direction, 1], [1, tileins.nearest_hole_direction, 1], [1, tileins.what_exists_front, 1], [1, tileins.what_exists_right, 1], 
[1, tileins.what_exists_left, 1], [1, tileins.what_exists_back, 1], [1, tileins.nearest_tile_direction, 1], 
[1, tileins.second_nearest_tile_direction, 1], [1, tileins.nearest_hole_direction, 1], [2, tileins.move_forward, 5], 
[2, tileins.turn_right, 5], [2, tileins.turn_left, 5], [2, tileins.stay, 5], [2, tileins.move_forward, 5], 
[2, tileins.turn_right, 5], [2, tileins.turn_left, 5], [2, tileins.stay, 5], [2, tileins.move_forward, 5], 
[2, tileins.turn_right, 5], [2, tileins.turn_left, 5], [2, tileins.stay, 5], [2, tileins.move_forward, 5], 
[2, tileins.turn_right, 5], [2, tileins.turn_left, 5], [2, tileins.stay, 5], [2, tileins.move_forward, 5], 
[2, tileins.turn_right, 5], [2, tileins.turn_left, 5], [2, tileins.stay, 5]]

# Node range definitions for GNP network
beginning_judgement_node = 2
end_judgement_node = 36
beginning_processing_node = 37
end_processing_node = 56

# Alternative smaller network configuration (PG=1)
# base_individual = [[], [0, 0, 0], [1, tileins.what_exists_front, 1], [1, tileins.what_exists_right, 1], 
# [1, tileins.what_exists_left, 1], [1, tileins.what_exists_back, 1], [1, tileins.nearest_tile_direction, 1], 
# [1, tileins.second_nearest_tile_direction, 1], [1, tileins.nearest_hole_direction, 1], [2, tileins.move_forward, 5], 
# [2, tileins.turn_right, 5], [2, tileins.turn_left, 5], [2, tileins.stay, 5]]

# Alternative node ranges (PG=1)
# beginning_judgement_node = 2
# end_judgement_node = 8
# beginning_processing_node = 9
# end_processing_node = 12


# ============================================================================
# SIMULATION CONFIGURATION
# ============================================================================

initial_tiles_index = variables.initial_tiles_index
initial_holes_index = variables.initial_holes_index
initial_agents_index = variables.initial_agents_index
initial_remaining_step = variables.initial_remaining_step
delay_time = variables.delay_time
each_step_according_to_di = variables.each_step_according_to_di


# ============================================================================
# GENETIC ALGORITHM PARAMETERS
# ============================================================================

population_size = variables.population_size # should be even
pc = variables.pc # probability of crossover
pm = variables.pm # probability of mutation
fitness_bias = variables.fitness_bias # bias for calculate in roullete wheel
epoch_number = variables.epoch_number
run_GA_times = variables.run_GA_times


# ============================================================================
# CLASSES
# ============================================================================

class object_head:
    """objects that have angle (like agent or tile)"""
    index = []
    head = 0
    def __init__(self, index, head):
        self.index = index
        self.head = head
    def heading(self):
        return self.head
    def setheading(self, head):
        self.head = head


# ============================================================================
# POPULATION INITIALIZATION
# ============================================================================

def generate_connection_gene(individual, connection_genes_number, begin_node, end_node):
    """Generate random connection genes for nodes in the specified range."""
    for i in range(begin_node, end_node): # loop for each node
        for j in range(connection_genes_number): # judgement nodes and processing nodes have differente connection genes number
            candidate_node = i # because connection genes should not include current node (to avoid loop in network)
            while candidate_node == i:
                candidate_node = random.randrange(2, len(individual))
            individual[i].append(candidate_node)
            individual[i].append(0)
    return individual


def initialize_population(individuals, population_size):
    """Initialize population with fixed node genes and random connection genes."""
    for i in range(population_size): # loop for each individual
        individuals.append([])
        for j in range(len(initial_agents_index)): # create one GNP network per agent
            individuals[i].append(copy.deepcopy(base_individual))
            # start node
            individuals[i][j] = generate_connection_gene(individuals[i][j], 1, 1, 2)
            # judgement node
            individuals[i][j] = generate_connection_gene(individuals[i][j], 4, beginning_judgement_node, end_judgement_node + 1)
            # processing node
            individuals[i][j] = generate_connection_gene(individuals[i][j], 1, beginning_processing_node, end_processing_node + 1)


# ============================================================================
# FITNESS CALCULATION
# ============================================================================

def nearest_hole_distance(tile, holes):
    """Calculate the minimum distance from a tile to any hole."""
    distances = []
    for hole in holes:
        distances.append(comins.distance_between(tile, hole))
    return min(distances)


def calculate_distance_between_tiles_and_holes(tiles, holes):
    """for use in calculate_fitness()"""
    summation = 0
    for tile in tiles:
        summation += nearest_hole_distance(tile, holes)
    return summation


def get_dropped_tile(tiles, holes):
    """Count how many tiles have been dropped into holes."""
    summation = 0
    for tile in tiles:
        for hole in holes:
            if tile.index == hole.index:
                summation += 1
    return summation


def calculate_fitness(tiles, holes, distance_between_tiles_and_holes_at_start, rest_steps):
    """Calculate fitness based on dropped tiles, distance improvement, and remaining steps."""
    approached_distance = distance_between_tiles_and_holes_at_start - calculate_distance_between_tiles_and_holes(tiles, holes)
    rest_step = round(sum(rest_steps) / len(rest_steps))
    return 100 * get_dropped_tile(tiles, holes) + round(20 * approached_distance) + rest_step


# ============================================================================
# AGENT EXECUTION (GNP NETWORK TRAVERSAL)
# ============================================================================

def run_agent(individual, agent, agent_index, rest_steps, agents, tiles, holes, node):
    """Execute one step for a single agent using its own GNP network structure."""
    temp_for_step = 0 # to use for calculate rest step
    while temp_for_step < each_step_according_to_di:
        if individual[node][0] == 1: # judgement node
            result = individual[node][1](agent, agents, tiles, holes)
            temp_for_step += individual[node][2]
        elif individual[node][0] == 2: # processing node
            individual[node][1](agent, agents, tiles, holes)
            result = 0 # because all processing node have one connection gene
            temp_for_step += individual[node][2]
        node = individual[node][3 + result * 2]
    # print(tiles[0].index)
    rest_steps[agent_index] -= 1
    return node


def run_individual(individual, agents, tiles, holes):
    """run agents with individual GNP network"""
    rest_steps = []
    # print(tiles[0].index)
    for i in range(len(agents)):
        rest_steps.append(initial_remaining_step) # rest step for each agent
    distance_between_tiles_and_holes_at_start = calculate_distance_between_tiles_and_holes(tiles, holes)
    nodes_for_each_agent = []
    for i in range(len(agents)):
        nodes_for_each_agent.append(individual[i][1][3]) # each agent starts at its own network's start node
        # nodes_for_each_agent = [individual[1][3], individual[1][3], individual[1][3]]
    for i in range(initial_remaining_step):
        for j in range(len(agents)):
            nodes_for_each_agent[j] = run_agent(individual[j], agents[j], j, rest_steps, agents, tiles, holes, nodes_for_each_agent[j])
            if calculate_distance_between_tiles_and_holes(tiles, holes) == 0:
                return calculate_fitness(tiles, holes, distance_between_tiles_and_holes_at_start, rest_steps)
    
    return calculate_fitness(tiles, holes, distance_between_tiles_and_holes_at_start, rest_steps)


def fitness(individual):
    """Calculate fitness for an individual by running it on a fresh simulation."""
    agents = []
    for i in range(len(initial_agents_index)):
        agents.append(object_head(initial_agents_index[i], 0))
    tiles = []
    for i in range(len(initial_tiles_index)):
        tiles.append(object_head(initial_tiles_index[i], 0))
    holes = []
    for i in range(len(initial_holes_index)):
        holes.append(object_head(initial_holes_index[i], 0))

    return run_individual(individual, agents, tiles, holes)


# ============================================================================
# SELECTION METHODS
# ============================================================================

def roullete_wheel(fitnesses):
    """Select an individual using roulette wheel selection."""
    # fitnesses = [-9, -6, -1, -14, -1, -9, 9, -1, -1, -1, -6, -1, -1, -1, -1, -1, -1, -9, -1, 299]
    min_of_fitness = min(fitnesses)
    for i in range(len(fitnesses)):
        fitnesses[i] = fitnesses[i] - min_of_fitness + fitness_bias
    sum_of_fitnesses = sum(fitnesses)
    rand_value = random.random()
    temp = 0
    for i in range(len(fitnesses)):
        temp += fitnesses[i] / sum_of_fitnesses
        if rand_value < temp:
            return i


def selection(fitnesses):
    """Select two parents using roulette wheel selection."""
    selected_parents = []
    for i in range(2):
        selected_parents.append(roullete_wheel(fitnesses.copy()))
    return selected_parents


# ============================================================================
# GENETIC OPERATORS
# ============================================================================

def crossover(individual1, individual2):
    """Perform standard crossover between two GNP networks."""
    for i in range(len(individual1)):
        if random.random() < pc:
            temp = individual1[i][3:]
            individual1[i][3:] = individual2[i][3:]
            individual2[i][3:] = temp
    
    return [individual1, individual2]


def mutation(individual):
    """Apply standard mutation to a GNP network."""
    for i in range(len(individual)):
        for j in range(3, len(individual[i]), 2):
            if random.random() < pm:
                candidate_node = i # because connection genes should not include current node (to avoid loop in network)
                while candidate_node == i:
                    candidate_node = random.randrange(2, len(individual))
                individual[i][j] = candidate_node


# ============================================================================
# FILE I/O
# ============================================================================

def convert_individual_to_GA_format(individual):
    """convert individual from GA functions to functions that write for use in tkinter"""
    for i in range(1, len(individual)):
        for j in range(3, len(individual[i])):
            base_individual[i].append(individual[i][j])
    return base_individual


def save_to_file(individual):
    """Save GNP individual to result.txt file."""
    try:
        f = open("result.txt", "a")
    except:
        print('interrupt in save_to_file function')
    individual_string = str(individual)
    result = ''
    for i in range(len(individual_string)):
        if individual_string[i] != '<' and individual_string[i] != '>':
            result += individual_string[i]
        else:
            result += '"'
    try:
        f.write(result + "\n\n\n")
        f.close()
    except:
        print('interrupt in save_to_file function')


# ============================================================================
# MAIN GENETIC ALGORITHM FUNCTIONS
# ============================================================================

def get_average(fitnesses_of_all):
    """get average of each epoch in all fitnesses(as the number of run_GA_times"""
    result = []
    for i in range(epoch_number):
        sum_of_each_epoch_number = 0
        result.append([])
        for j in range(run_GA_times):
            sum_of_each_epoch_number += fitnesses_of_all[j][i]
        result[i] = sum_of_each_epoch_number // run_GA_times
    return result


def save_to_excel(fitnesses):
    """Save fitness data to Excel file."""
    excel_file_name = 'fitnesses.xlsx'
    workbook = xl.Workbook()
    workbook.save(excel_file_name)
    sheet_name = 'Sheet'
    excel_file = xl.load_workbook(excel_file_name)
    excel_sheet = excel_file[sheet_name]
    for i in range(len(fitnesses)):
        try:
            excel_sheet.cell(i + 1, 1).value = fitnesses[i]
            excel_file.save(excel_file_name)
        except PermissionError:
            print('permission denied!')


def save_to_excel_complete_report(fitnesses_of_all):
    """Save complete fitness report to Excel file."""
    excel_file_name = 'fitnesses_complete_report.xlsx'
    
    # Create a new workbook
    workbook = xl.Workbook()
    sheet_name = 'Fitness Data'
    
    # Access the active sheet (automatically created)
    excel_sheet = workbook.active
    excel_sheet.title = sheet_name
    
    for i in range(len(fitnesses_of_all)):
        for j in range(epoch_number):
            excel_sheet.cell(row=j+1, column=i+1).value = fitnesses_of_all[i][j]
    
    try:
        workbook.save(excel_file_name)
        print("Data saved to", excel_file_name)
    except PermissionError:
        print('Permission denied! Unable to save excel file.')


def run_GA(max_fitness_of_each_epoch):
    """Run the genetic algorithm for evolving GNP networks."""
    individuals = []
    initialize_population(individuals, population_size) # fill individuals with fixed node genes and random connection genes
    fitnesses = []
    for i in range(population_size):
        fitnesses.append(fitness(individuals[i]))
    
    for i in range(epoch_number):
        print('epoch number : ' + str(i))
        print(fitnesses)
        fitnesses = []
        for j in range(population_size):
            fitnesses.append(fitness(individuals[j]))
        new_individuals = [] # for next epochs
        individual_with_max_fitness = individuals[fitnesses.index(max(fitnesses))] # to hold 2 best
        new_individuals.append(copy.deepcopy(individual_with_max_fitness)) # to hold 2 best
        new_individuals.append(copy.deepcopy(individual_with_max_fitness)) # to hold 2 best
        max_fitness_of_each_epoch.append(max(fitnesses))
        
        for j in range((population_size // 2) - 1): # to hold 2 best
            selected_individual_index = selection(fitnesses)
            new_individuals.append([])
            new_individuals.append([])
            # Perform crossover on each agent's network separately
            for k in range(len(initial_agents_index)):
                crossovered = crossover(
                    copy.deepcopy(individuals[selected_individual_index[0]][k]),
                    copy.deepcopy(individuals[selected_individual_index[1]][k])
                )
                new_individuals[-2].append(crossovered[0])
                new_individuals[-1].append(crossovered[1])
        
        individuals = new_individuals
        # Apply mutation to each agent's network separately
        for j in range(2, len(new_individuals)): # hold 2 best
            for k in range(len(initial_agents_index)):
                mutation(new_individuals[j][k])

    fitnesses = []
    for j in range(population_size):
        fitnesses.append(fitness(individuals[j]))
    print(fitnesses)
    print(max(fitnesses))
    print(individuals[fitnesses.index(max(fitnesses))])
    max_fitness_of_each_epoch.append(max(fitnesses))
    max_fitness_of_each_epoch.pop(0) # pop first element (because before this line len(max_fitness_of_each_epoch) is epoch_number + 1)
    save_to_file(individuals[fitnesses.index(max(fitnesses))])
    return individuals[fitnesses.index(max(fitnesses))]


def run_GA_concurrent():
    """Run multiple GA instances concurrently using threads."""
    time_at_start = time.time()
    threads = []
    fitnesses_of_all = []

    for i in range(run_GA_times):
        fitnesses_of_all.append([])
        threads.append(threading.Thread(target=run_GA, args=(fitnesses_of_all[i], )))
    for i in range(run_GA_times):
        threads[i].start()
    for i in range(run_GA_times):
        threads[i].join()
    print(fitnesses_of_all)
    print('time:' + str(time.time() - time_at_start))
    print(get_average(fitnesses_of_all))
    save_to_excel(get_average(fitnesses_of_all))
    save_to_excel_complete_report(fitnesses_of_all)
    a = input('finish! press enter to exit ...')


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

if __name__ == '__main__':
    run_GA_concurrent()